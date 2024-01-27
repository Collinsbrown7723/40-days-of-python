import openpyxl

wb = openpyxl.load_workbook("store.xlsx")

sheet1,sheet2 = wb.sheetnames

sheet = wb[sheet1]
sheet.title = "Products for sale"
wb.create_sheet("another sheet")
wb.create_sheet("first sheet",0)
sheet3 = wb["another sheet"]
wb.remove(sheet3)

tobecopied = wb["first sheet"]

destination = wb.copy_worksheet(tobecopied)
print(destination.title)
print(wb.sheetnames)
print(sheet.title)