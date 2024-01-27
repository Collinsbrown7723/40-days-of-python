import openpyxl
wb = openpyxl.load_workbook("store.xlsx",data_only=True)

print(type(wb))
print(wb.sheetnames)

for sheet in wb:
    print(sheet)

sheet = wb["Products"]
#sheet = wb.active
data1 = sheet["b2"]
data2 = sheet["c2"]

print(data1.value,data2.value)
print(sheet.cell(row=4,column=2).value)
print(data1.row,data1.column)
print(sheet["d4"].parent)
print(dir(data1))

cell_range = sheet["B2:C11"]

# for product,price in cell_range:
#     price(f"product = {product.value} price = {price.value}")
print(sheet.dimensions)
for a,b,c,d,e in sheet[sheet.dimensions]:
    print(a.value,b.value,c.value,d.value,e.value)
    
for row in sheet.rows:
    for cell in row:
        print(f"{cell.value} ",end="")
for row in sheet.values:
    print(row)
        
    