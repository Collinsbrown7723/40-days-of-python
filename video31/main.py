import openpyxl

wb = openpyxl.Workbook()

sheet1 = wb.create_sheet("company")

sheet1["A1"] = "year"
sheet1["B1"] = "sales"

cell_range = []

data_collected = {"2017":"150000","2018":"180000","2019":"210000","2020":"125000"}


years = list()
sales = list()
for a,b in data_collected.items():
    years.append(a)
    sales.append(b)
i = 0
for y,k in sheet1["A2:B5"]:
    y.value = years[i]
    k.value = sales[i]
    print(y,k,i)
    i+=1


"""Challenge #2

Change the solution from the previous challenge and set the sheet title to COMPANY SALES."""

sheet1.title = "COMPANY SALES"
print(wb.sheetnames)
wb.save("Company.xlsx")

